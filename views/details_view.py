import tkinter as tk


from tkinter import ttk
from tkinter import *

# choose file
from tkinter import filedialog

# warning box for choose-side
from tkinter import messagebox

# get relpath
import os

# stitch images
import sys
from PIL import Image

# date entry
# from tkcalendar import Calendar, DateEntry
from tkcalendar import DateEntry

# update prefs
import json

# append2excel and master export
from openpyxl import load_workbook
import pandas as pd
from pathlib import Path


# multithreading
import threading
import queue


class DETAILS_View(tk.Frame):
	def __init__(self, parent, controller, master_dict):
		tk.Frame.__init__(self, parent)
		self.controller = controller

		self.master_dict = master_dict

		
		# to avoid iteration bugs
		self.med_image = ""

		self.pat_deets = PatDetailsFrame(self)		
		self.pat_deets.pack(side="left", anchor="nw") # stick to the top
		
		self.pat_mments = PatMeasurementsFrame(self)		
		self.pat_mments.pack(side="left", anchor="nw", padx=(30,0)) # stick to the top




		
		# # for stitching images
		# self.im1 = ""
		# self.im2 = ""
		# self.im_result_name = ""


		# tk.Label(self, text="POINT SIZES", font=("TkDefaultFont", 12)).grid(sticky="W", column=0, row=7, pady=(60, 10), padx=(10,0))

		
		# self.pre_scanno_var		= StringVar()
		# self.pre_ap_var	 		= StringVar()
		# self.pre_lat_var	 	= StringVar()
		# self.pre_sky_var		= StringVar()
		# self.post_scanno_var	= StringVar()
		# self.post_ap_var	 	= StringVar()
		# self.post_lat_var		= StringVar()
		# self.post_sky_var		= StringVar()



		# tk.Label(self, text="PRE-SCANNO").grid(sticky="W", column=0, row=8, padx=(10,0))
		# sp1 = Spinbox(self, textvariable=self.pre_scanno_var, from_= 1, to = 30, width=10).grid(column=1, row=8)

		# tk.Label(self, text="PRE-AP").grid(sticky="W", column=0, row=9, padx=(10,0))
		# sp2 = Spinbox(self, textvariable=self.pre_ap_var, from_= 1, to = 30, width=10).grid(column=1, row=9)

		# tk.Label(self, text="PRE-LAT").grid(sticky="W", column=0, row=10, padx=(10,0))
		# sp3 = Spinbox(self, textvariable=self.pre_lat_var, from_= 1, to = 30, width=10).grid(column=1, row=10)

		# tk.Label(self, text="PRE-SKY").grid(sticky="W", column=0, row=11, padx=(10,0))
		# sp4 = Spinbox(self, textvariable=self.pre_sky_var, from_= 1, to = 30, width=10).grid(column=1, row=11)

		# tk.Label(self, text="POST-SCANNO").grid(sticky="W", column=0, row=12, padx=(10,0), pady=(20, 0))
		# sp5 = Spinbox(self, textvariable=self.post_scanno_var, from_= 1, to = 30, width=10).grid(column=1, row=12, pady=(20, 0))

		# tk.Label(self, text="POST-AP").grid(sticky="W", column=0, row=13, padx=(10,0))
		# sp6 = Spinbox(self, textvariable=self.post_ap_var, from_= 1, to = 30, width=10).grid(column=1, row=13)

		# tk.Label(self, text="POST-LAT").grid(sticky="W", column=0, row=14, padx=(10,0))
		# sp7 = Spinbox(self, textvariable=self.post_lat_var, from_= 1, to = 30, width=10).grid(column=1, row=14)

		# tk.Label(self, text="POST-SKY").grid(sticky="W", column=0, row=15, padx=(10,0))
		# sp8 = Spinbox(self, textvariable=self.post_sky_var, from_= 1, to = 30, width=10).grid(column=1, row=15)

		# # tk.Label(self, text="v0.4").grid(sticky="W", row=16, pady=(30,0), padx=(10,0))
		# tk.Label(self, text=self.controller.app_version).grid(sticky="W", row=16, pady=(30,0), padx=(10,0))




		# tk.Label(self, text="IMAGE STITCH", font=("TkDefaultFont", 12)).grid(sticky="W", column=3, row=0, pady=(30, 10), padx=(80,0), columnspan=3)
		
		# btn_set_im1 = ttk.Button(self, text="IMG 1", width=10, command=lambda: self.set_stitch_images("im1"))
		# btn_set_im1.grid(sticky="W", column=3, row=3, padx=(80,0))
		# btn_set_im2 = ttk.Button(self, text="IMG 2", width=10, command=lambda: self.set_stitch_images("im2"))
		# btn_set_im2.grid(sticky="W", column=4, row=3)
		# btn_stitch 	= ttk.Button(self, text="GO", width=10, command=lambda: self.set_stitch_images("go"))
		# btn_stitch.grid(sticky="W", column=5, row=3)


		# btn_rot_im1 = ttk.Button(self, text="ROT 1", width=10, command=lambda: self.rotateStichImage("im1"))
		# btn_rot_im1.grid(sticky="W", column=3, row=4, padx=(80,0))
		# btn_rot_im2 = ttk.Button(self, text="ROT 2", width=10, command=lambda: self.rotateStichImage("im2"))
		# btn_rot_im2.grid(sticky="W", column=4, row=4)



		# self.label1 = tk.Label(self, text="Img Name:").grid(sticky="W", column=3, row=1, padx=(80,0))
		# self.img_stitch_name = tk.Entry(self)
		# self.img_stitch_name.grid(sticky="W", column=3, row=2, padx=(80,0), columnspan=3)




	def updateTableValues(self):
		print('updateTableValues')		

		self.pat_mments.updateTableFromDict()


	def callback(self, P):
		if str.isdigit(P) or P == "":
			return True
		else:
			return False

	def callback2(self, sv):
		print(sv.get())

	def is_set_med_image(self):
		# This class has no image related to it
		# exception for DETAILS_View
		return True

	def update_dict(self, master_dict):		
		# update dictionaries
		self.master_dict = master_dict

		self.add_details_masterdict()
		
		self.pat_mments.update_dict(master_dict)

		self.pat_deets.update_dict(master_dict)

		
		# self.updateSpinboxes()



	def add_details_masterdict(self):
		
		if "DETAILS" not in self.master_dict.keys():		# details dont exist new patient
			# print("details reached")
			self.master_dict["DETAILS"] = 	{
										"F_NAME":None,
										"L_NAME":None,
										"AGE":None,
										"SEX":None,
										"R_KL_GRADE": None,
										"L_KL_GRADE": None,
										"R_SX_DATE": None,
										"L_SX_DATE": None,
										"R_PROS":None,
										"L_PROS":None
									}
			self.controller.save_json()


	def updateSpinboxes(self):
		self.pre_scanno_var.set(self.master_dict["POINT_SIZES"]["PRE-SCANNO"])
		self.pre_ap_var.set(self.master_dict["POINT_SIZES"]["PRE-AP"])
		self.pre_lat_var.set(self.master_dict["POINT_SIZES"]["PRE-LAT"])
		self.pre_sky_var.set(self.master_dict["POINT_SIZES"]["PRE-SKY"])
		self.post_scanno_var.set(self.master_dict["POINT_SIZES"]["POST-SCANNO"])
		self.post_ap_var.set(self.master_dict["POINT_SIZES"]["POST-AP"])
		self.post_lat_var.set(self.master_dict["POINT_SIZES"]["POST-LAT"])
		self.post_sky_var.set(self.master_dict["POINT_SIZES"]["POST-SKY"])

		if self.master_dict["DETAILS"]["F_NAME"] != None:
			self.sv_fname.set(self.master_dict["DETAILS"]["F_NAME"])
		if self.master_dict["DETAILS"]["L_NAME"] != None:
			self.sv_lname.set(self.master_dict["DETAILS"]["L_NAME"])
		if self.master_dict["DETAILS"]["AGE"] != None:
			self.sv_age.set(self.master_dict["DETAILS"]["AGE"])

		if self.master_dict["DETAILS"]["SEX"] != None:
			self.sex.set(self.master_dict["DETAILS"]["SEX"])

		# self.pre_scanno_var.trace("w", lambda name, index, mode, dict_key="PRE-SCANNO", var=self.pre_scanno_var: self.update_pointsize_dict(dict_key, var))
		# self.post_scanno_var.trace("w", lambda name, index, mode, dict_key="POST-SCANNO", var=self.post_scanno_var: self.update_pointsize_dict(dict_key, var))

		self.pre_scanno_var.trace("w", lambda name, index, mode, dict_key="PRE-SCANNO", var=self.pre_scanno_var: self.update_pointsize_dict(dict_key, var))
		self.pre_ap_var.trace("w", lambda name, index, mode, dict_key="PRE-AP", var=self.pre_ap_var: self.update_pointsize_dict(dict_key, var))
		self.pre_lat_var.trace("w", lambda name, index, mode, dict_key="PRE-LAT", var=self.pre_lat_var: self.update_pointsize_dict(dict_key, var))
		self.pre_sky_var.trace("w", lambda name, index, mode, dict_key="PRE-SKY", var=self.pre_sky_var: self.update_pointsize_dict(dict_key, var))
		self.post_scanno_var.trace("w", lambda name, index, mode, dict_key="POST-SCANNO", var=self.post_scanno_var: self.update_pointsize_dict(dict_key, var))
		self.post_ap_var.trace("w", lambda name, index, mode, dict_key="POST-AP", var=self.post_ap_var: self.update_pointsize_dict(dict_key, var))
		self.post_lat_var.trace("w", lambda name, index, mode, dict_key="POST-LAT", var=self.post_lat_var: self.update_pointsize_dict(dict_key, var))
		self.post_sky_var.trace("w", lambda name, index, mode, dict_key="POST-SKY", var=self.post_sky_var: self.update_pointsize_dict(dict_key, var))


		# trace entrys
		self.sv_fname.trace_add("write", lambda name, index, mode, entry_type="F-NAME", var=self.sv_fname: self.entry_callback(entry_type, self.sv_fname))
		self.sv_lname.trace_add("write", lambda name, index, mode, entry_type="L-NAME", var=self.sv_lname: self.entry_callback(entry_type, self.sv_lname))
		self.sv_age.trace_add("write", lambda name, index, mode, entry_type="AGE", var=self.sv_age: self.entry_callback(entry_type, self.sv_age))


		self.sex.trace_add("write", lambda name, index, mode, var=self.sex: self.radio_callback(self.sex))



	def update_pointsize_dict(self, dict_key, var):
		print('{} key update {}'.format(dict_key, var.get()))

		# no blank vals
		if var.get() != "":
			# try int cast
			try: 
				# make sure positive
				# and in range of 5 - 30
				int_check = int(var.get())
				# int_check = float(var.get())
				if int_check > 0 and int_check < 31:
					self.master_dict["POINT_SIZES"][dict_key] = int_check

					# pass the dictkey which is also the view(page_name) which updates all objects with new point size
					self.controller.redrawNewPointSize(dict_key)
					
					# save the new point size to disk
					self.controller.save_json()

			# int cast failed, must be text entry
			except Exception as e:
				print(e)
				# raise e



	def entry_callback(self, entry_type, var):
		print('{} entry_type update {}'.format(entry_type, var.get()))


		val = var.get().strip()
		if val != "":
			if entry_type == "F-NAME":
				self.master_dict["DETAILS"]["F_NAME"] = val
				self.controller.save_json()
			elif entry_type == "L-NAME":
				self.master_dict["DETAILS"]["L_NAME"] = val
				self.controller.save_json()
			elif entry_type == "AGE":
				self.master_dict["DETAILS"]["AGE"] = val
				self.controller.save_json()


	def radio_callback(self, var):

		sex = var.get().strip()

		if sex != None:
			self.master_dict["DETAILS"]["SEX"] = sex
			self.controller.save_json()

		print('radio callback {}'.format(sex))



	def set_stitch_images(self, btn_type):
		if btn_type == "im1" or btn_type == "im2":
			print("set images")


			# get image
			image = filedialog.askopenfilename(initialdir=self.controller.working_dir)

			if isinstance(image, str) and image != "":

				dir_name = os.path.dirname(image)
				rel_path = os.path.relpath(image, dir_name)

				# only allow images from the working dir
				if self.controller.working_dir != dir_name:
					print("mis-match")
					self.warningBox("Image not from working directory")
					return

				if btn_type == "im1":
					self.im1 = rel_path
				elif btn_type == "im2":
					self.im2 = rel_path

		if btn_type == "go":
			# check if images are set
			if self.im1 == "":
				self.warningBox("Image 1 Not Selected")
				return
			if self.im2 == "":
				self.warningBox("Image 2 Not Selected")
				return

			# check if result image name is set
			im_name = self.img_stitch_name.get().strip()
			im_name = self.controller.working_dir + "/" + im_name + '.jpg'
			if im_name == "":
				self.warningBox("Stitched Image-name not set")
				return


			images = [Image.open(x) for x in [(self.controller.working_dir + "/" + self.im1), (self.controller.working_dir + "/" + self.im2)]]
			# images = [Image.open(x) for x in ['post_ap.jpg','post_lat.jpg']]

			
			widths, heights = zip(*(i.size for i in images))

			total_width = sum(widths)
			max_height = max(heights)

			new_im = Image.new('RGB', (total_width, max_height))

			x_offset = 0
			for im in images:
				new_im.paste(im, (x_offset,0))
				x_offset += im.size[0]

			new_im.save(im_name)
			self.im1 = ""
			self.im2 = ""
			self.img_stitch_name.delete(0, tk.END)
			self.warningBox("Success")


	def rotateStichImage(self, btn_type):
		im_path = ""
		if btn_type == "im1":
			if self.im1 == "":
				self.warningBox("Image 1 Not Selected")
				return
			im_path = self.controller.working_dir + "/" + self.im1

		elif btn_type == "im2":
			if self.im2 == "":
				self.warningBox("Image 2 Not Selected")
				return
			im_path = self.controller.working_dir + "/" + self.im2

		if im_path != "":
			img_rt_90 = self.rotate_img(im_path, 90)
			img_rt_90.save(im_path)
			self.warningBox("Rotated 90")


	def rotate_img(self, img_path, rt_degr):
		img = Image.open(img_path)
		return img.rotate(rt_degr, expand=1)



	def warningBox(self, message):
		'''Display a warning box with message'''
		messagebox.showwarning("Warning", message)


	def draw_presentation_images(self):
		print('draw_presentation_images')
		# pass
		self.controller.app_draw_pil()


	def details_tryPsFile(self):
		print('details_tryPsFile')
		# get into the view to access canvas
		self.controller.app_tryPsFile()



	def escapeFunc(self):
		pass
	def keySetRight(self):
		pass
	def keySetLeft(self):
		pass

				
				


class PatMeasurementsFrame(ttk.Frame):
	def __init__(self, parent, *args, **kwargs):
		tk.Frame.__init__(self, parent, *args, **kwargs)
		self.parent = parent

		self.master_dict = {}
		# tk.Label(self, text="PATIENT MEASUREMENTS", font=("TkDefaultFont", 12)).grid(sticky="W", column=0, row=0, columnspan=5)

		# self.frame_scano = tk.Frame(self, width=300,bg="#000")		
		# self.frame_ap	= tk.Frame(self, width=300,bg="#666666")
		# self.frame_lat	= tk.Frame(self, width=300,bg="yellow")
		# self.frame_sky	= tk.Frame(self, width=300,bg="red")

		self.frame_scano = tk.Frame(self, width=300, padx=20,pady=20)
		self.frame_ap	= tk.Frame(self, width=30, padx=30,pady=20)

		# frame_scano.pack(side='top', anchor="nw")
		# frame_ap.pack(side='top', anchor="ne")
		# frame_lat.pack(side='bottom', anchor="sw")
		# frame_sky.pack(side='bottom', anchor="se")

		self.frame_scano.grid(row=0,column=0, rowspan=3)
		self.frame_ap.grid(row=0,column=1, sticky="N")		

		self.table_trace_dict = {}


		self.populateFrame(self.frame_scano, ["HKA","MAD","MNSA","VCA","aFTA","mLDFA","aLDFA","EADFA","EADFPS","EADFDS","JDA","TAMD","MPTA","KJLO","KAOL","EADTA","EADTPS","EADTDS"])
		self.populateFrame(self.frame_ap,["FFLEX","PCOR","ACOR","TSLOPE","ISR","SA","PTILT","PPBA","FVAR/VAL","TVAR/VAL","FFLE/EXT","TFLE/EXT","JCA","LPFA","MPFA","LDTA","ANKLE_SLOPE","pMA"])		

		


	def update_dict(self, master_dict):		
		self.master_dict = master_dict
		self.updateTableFromDict()


	def populateFrame(self, frame, label_list):

		tk.Label(frame, text="PRE-OP").grid(sticky="WE", row=0, column=1, columnspan=2)
		tk.Label(frame, text="POST-OP").grid(sticky="WE", row=0, column=3, columnspan=2)
		tk.Label(frame, text="RIGHT").grid(sticky="E", row=1, column=1)
		tk.Label(frame, text="LEFT").grid(sticky="E", row=1, column=2)
		tk.Label(frame, text="RIGHT").grid(sticky="E", row=1, column=3)
		tk.Label(frame, text="LEFT").grid(sticky="E", row=1, column=4)
				

		rownum = 2
		colwidth = 80

		bg = True

		for label in label_list:

			sv_preR 	= StringVar()
			sv_preL 	= StringVar()
			sv_postR 	= StringVar()
			sv_postL 	= StringVar()

			# add vars to dict for easy update later
			self.table_trace_dict[label] = [sv_preR,sv_preL,sv_postR,sv_postL]
			
			lab0 = tk.Label(frame, text=label, width=9, anchor='w', padx=10, pady=4)
						
			lab1 = tk.Label(frame, textvariable=sv_preR, width=9, anchor='e', padx=10, pady=4)			
			lab2 = tk.Label(frame, textvariable=sv_preL, width=9, anchor='e', padx=10, pady=4)
			lab3 = tk.Label(frame, textvariable=sv_postR, width=9, anchor='e', padx=10, pady=4)
			lab4 = tk.Label(frame, textvariable=sv_postL, width=9, anchor='e', padx=10, pady=4)

			lab0.grid(sticky="W", row=rownum, column=0)
			lab1.grid(sticky="E", row=rownum, column=1)
			lab2.grid(sticky="E", row=rownum, column=2)
			lab3.grid(sticky="E", row=rownum, column=3)
			lab4.grid(sticky="E", row=rownum, column=4)
			
			

			# preR 	= self.master_dict["EXCEL"]["PRE-OP"]["RIGHT"][label]
			# preL 	= self.master_dict["EXCEL"]["PRE-OP"]["LEFT"][label]
			# postR 	= self.master_dict["EXCEL"]["POST-OP"]["RIGHT"][label]
			# postL 	= self.master_dict["EXCEL"]["POST-OP"]["LEFT"][label]

			# if preR == None: preR = "" 
			# if preL == None: preL = "" 
			# if postR == None: postR = "" 
			# if postL == None: postL = "" 

			# l0 = tk.Label(frame, text=label, width=9, anchor='w', padx=10, pady=4)
			# l0.grid(sticky="W", row=rownum, column=0)
			# l1 = tk.Label(frame, text=preR, width=9, anchor='e', padx=10, pady=4)
			# l1.grid(sticky="E", row=rownum, column=1)
			# l2 = tk.Label(frame, text=preL, width=9, anchor='e', padx=10, pady=4)
			# l2.grid(sticky="E", row=rownum, column=2)
			# l3 = tk.Label(frame, text=postR, width=9, anchor='e', padx=10, pady=4)
			# l3.grid(sticky="E", row=rownum, column=3)
			# l4 = tk.Label(frame, text=postL, width=9, anchor='e', padx=10, pady=4)
			# l4.grid(sticky="E", row=rownum, column=4)

			if bg:
				lab0['bg'] = '#ffffff'
				lab1['bg'] = '#ffffff'
				lab2['bg'] = '#ffffff'
				lab3['bg'] = '#ffffff'
				lab4['bg'] = '#ffffff'
			bg = not bg

			
			frame.grid_columnconfigure(1, minsize=colwidth)
			frame.grid_columnconfigure(2, minsize=colwidth)
			frame.grid_columnconfigure(3, minsize=colwidth)
			frame.grid_columnconfigure(4, minsize=colwidth)

			rownum=rownum+1


	def updateTableFromDict(self):

		# on startup parsing json takes some time
		if not self.master_dict:			
			return

		refresh_list = ["HKA",
						"MAD",
						"MNSA",
						"VCA",
						"aFTA",
						"mLDFA",
						"aLDFA",
						"EADFA",
						"EADFPS",
						"EADFDS",
						"JDA",
						"TAMD",
						"MPTA",
						"KJLO",
						"KAOL",
						"EADTA",
						"EADTPS",
						"EADTDS",
						"FFLEX",
						"PCOR",
						"ACOR",
						"TSLOPE",
						"ISR",
						"SA",
						"PTILT",
						"PPBA",
						"MISCELL",
						"FVAR/VAL",
						"TVAR/VAL",
						"FFLE/EXT",
						"TFLE/EXT",
						"JCA",
						"LPFA",
						"MPFA",
						"LDTA",
						# "ANKLE_SLOPE",
						"pMA"]


		for label in refresh_list:
			
			try:
				preR = self.master_dict['EXCEL']['PRE-OP']['RIGHT'][label]
				preL = self.master_dict['EXCEL']['PRE-OP']['LEFT'][label]
				postR = self.master_dict['EXCEL']['POST-OP']['RIGHT'][label]
				postL = self.master_dict['EXCEL']['POST-OP']['LEFT'][label]

				if preR != None:
					self.table_trace_dict[label][0].set(preR)
				if preL != None:
					self.table_trace_dict[label][1].set(preL)

				if postR != None:
					self.table_trace_dict[label][2].set(postR)
				if postL != None:
					self.table_trace_dict[label][3].set(postL)
			except Exception as e:
				# raise e
				print(e)
				



class PatDetailsFrame(ttk.Frame):
	def __init__(self, parent, *args, **kwargs):
		tk.Frame.__init__(self, parent, *args, **kwargs)
		self.parent = parent

		self.master_dict = {}

		# self.upper_details = tk.Frame(self, height=100, width=300,bg="#000")		
		# self.lower_details	= tk.Frame(self, height=100, width=300,bg="#666666")
		# self.excel_frame	= tk.Frame(self, height=100, width=300,bg="yellow")
		# self.frame_sky	= tk.Frame(self, height=100, width=300,bg="red")

		self.upper_details = tk.Frame(self, height=100, width=300)
		self.lower_details	= tk.Frame(self, height=100, width=300)
		self.excel_frame	= tk.Frame(self, height=100, width=300)
		self.frame_sky	= tk.Frame(self, height=100, width=300)

		self.upper_details.grid(sticky="W", row=0, column=0, padx=20, pady=(20,0))
		self.lower_details.grid(sticky="W", row=1, column=0, padx=20, pady=(10,20))
		self.excel_frame.grid(sticky="W", row=2, column=0, padx=20, pady=20)
		self.frame_sky.grid(sticky="W", row=3, column=0, padx=20, pady=20)




		self.tkr_excel_file = None
		self.ukr_excel_file = None

		self.sv_fname 	= tk.StringVar()	# F-Name
		self.sv_lname 	= tk.StringVar()	# L-Name
		self.sv_age 	= tk.StringVar()	# AGE
		self.sv_sex 	= tk.StringVar()	# Sex Radiobutton

		self.sv_r_kl 	= tk.StringVar()	# K-L-Grade
		self.sv_l_kl 	= tk.StringVar()	# K-L-Grade

		self.sv_utkr 	= tk.StringVar()	# ukr/tkr radio

		self.sv_r_pros = tk.StringVar()		# prothesis drop-down
		self.sv_l_pros = tk.StringVar()		# prothesis drop-down

		# vars for date entry
		self.r_cal = None
		self.l_cal = None


		self.sv_fname.trace_add("write", lambda name, index, mode, entry_type="F-NAME", var=self.sv_fname: self.entry_callback(entry_type, self.sv_fname))
		self.sv_lname.trace_add("write", lambda name, index, mode, entry_type="L-NAME", var=self.sv_lname: self.entry_callback(entry_type, self.sv_lname))
		self.sv_age.trace_add("write", lambda name, index, mode, entry_type="AGE", var=self.sv_age: self.entry_callback(entry_type, self.sv_age))

		self.sv_r_kl.trace_add("write", lambda name, index, mode, entry_type="R-KL", var=self.sv_r_kl: self.entry_callback(entry_type, self.sv_r_kl))
		self.sv_l_kl.trace_add("write", lambda name, index, mode, entry_type="L-KL", var=self.sv_l_kl: self.entry_callback(entry_type, self.sv_l_kl))


		self.sv_sex.trace_add("write", lambda name, index, mode, var=self.sv_sex: self.radio_callback(self.sv_sex))
		self.sv_utkr.trace_add("write", lambda name, index, mode, var=self.sv_utkr: self.utkr_callback(self.sv_utkr))

		self.sv_r_pros.trace_add("write", lambda name, index, mode, entry_type="R-PROS", var=self.sv_r_pros: self.entry_callback(entry_type, self.sv_r_pros))
		self.sv_l_pros.trace_add("write", lambda name, index, mode, entry_type="L-PROS", var=self.sv_l_pros: self.entry_callback(entry_type, self.sv_l_pros))


		self.constructUpperDetails(self.upper_details)
		self.constructLowerDetails(self.lower_details)
		self.constructExcelFrame(self.excel_frame)


	def update_dict(self, master_dict):
		self.master_dict = master_dict

		# populate entry values
		fname 	= self.master_dict["DETAILS"]["F_NAME"]
		lname 	= self.master_dict["DETAILS"]["L_NAME"]
		age 	= self.master_dict["DETAILS"]["AGE"]
		sex 	= self.master_dict["DETAILS"]["SEX"]
		r_kl 	= self.master_dict["DETAILS"]["R_KL_GRADE"]
		l_kl 	= self.master_dict["DETAILS"]["L_KL_GRADE"]
		r_pros 	= self.master_dict["DETAILS"]["R_PROS"]
		l_pros 	= self.master_dict["DETAILS"]["L_PROS"]
		r_sx 	= self.master_dict["DETAILS"]["R_SX_DATE"]
		l_sx 	= self.master_dict["DETAILS"]["L_SX_DATE"]

		if fname != None:
			self.sv_fname.set(fname)
		if lname != None:
			self.sv_lname.set(lname)
		if age != None:
			self.sv_age.set(age)
		if sex != None:
			self.sv_sex.set(sex)

		if r_kl != None:
			self.sv_r_kl.set(r_kl)
		if l_kl != None:
			self.sv_l_kl.set(l_kl)

		if r_pros != None:
			self.sv_r_pros.set(r_pros)
		if l_pros != None:
			self.sv_l_pros.set(l_pros)


		if r_sx != None and self.r_cal != None:
			self.r_cal.set_date(r_sx)
		if l_sx != None and self.l_cal != None:
			self.l_cal.set_date(l_sx)


	def constructUpperDetails(self, frame):
	
		frame.grid_columnconfigure(0, minsize=80)

		tk.Label(frame, text="PATIENT DETAILS", font=("TkDefaultFont", 12)).grid(sticky="W", column=0, row=0, columnspan=2)


		tk.Label(frame, text="First Name").grid(sticky="W", row=1, column=0)
		tk.Label(frame, text="Last Name").grid(sticky="W", row=2, column=0)
		self.e_fname 	= ttk.Entry(frame, width=40, textvariable=self.sv_fname)
		self.e_lname	= ttk.Entry(frame, width=40, textvariable=self.sv_lname)
		self.e_fname.grid(sticky="W",row=1, column=1, padx=5, pady=5)
		self.e_lname.grid(sticky="W",row=2, column=1, padx=5, pady=5)


		tk.Label(frame, text="Age").grid(sticky="W", row=3, column=0)

		age_sex_frame = ttk.Frame(frame)
		age_sex_frame.grid(sticky="W",row=3, column=1)

		
		self.e_age = ttk.Entry(age_sex_frame, width=8, textvariable=self.sv_age)
		self.e_age.pack(side="left", padx=5)

		tk.Label(age_sex_frame, text="Sex").pack(side="left", padx=(60,0))
		rb1 = ttk.Radiobutton(age_sex_frame, text='M', value='M', variable=self.sv_sex)
		rb2 = ttk.Radiobutton(age_sex_frame, text='F', value='F', variable=self.sv_sex)
		rb1.pack(side="left", padx=(5,0))	
		rb2.pack(side="left")


		# self.e_age 		= tk.Entry(self, textvariable=self.sv_age)
		
		



		
		# m_btn = tk.Radiobutton(frame, text='M', variable=self.sex, value="M",tristatevalue="x")
		# # m_btn.grid(sticky="W", row=5, column=1)
		# m_btn.grid(sticky="W",row=4, column=1)
		# f_btn = tk.Radiobutton(frame, text='F', variable=self.sex, value="F",tristatevalue="x")
		# f_btn.grid(sticky="W",row=5, column=1)
		# # f_btn.grid(sticky="W", row=6, column=1)


	def constructLowerDetails(self, frame):
		tk.Label(frame, text="RIGHT").grid(sticky="WE", row=0, column=1)
		tk.Label(frame, text="LEFT").grid(sticky="WE", row=0, column=2)

		frame.grid_columnconfigure(0, minsize=80)



		tk.Label(frame, text="K-L GRADE").grid(sticky="W", row=1, column=0)

		self.r_klgrade = ttk.Combobox(frame, width = 12, textvariable = self.sv_r_kl, state="readonly")
		self.l_klgrade = ttk.Combobox(frame, width = 12, textvariable = self.sv_l_kl, state="readonly")

		print(self.parent.controller.user_pref_obj['IMPLANT_NAMES']['UKR'])

		self.r_klgrade['values'] = ('NA','1','2','3','4')
		self.l_klgrade['values'] = ('NA','1','2','3','4')

		# self.r_klgrade['values'] = self.parent.controller.user_pref_obj['IMPLANT_NAMES']['UKR']
		# self.r_klgrade 	= ttk.Entry(frame, width=15)
		# self.l_klgrade 	= ttk.Entry(frame, width=15)
		self.r_klgrade.grid(row=1, column=1)
		self.l_klgrade.grid(row=1, column=2)



		tk.Label(frame, text="SX DATE").grid(sticky="W", row=2, column=0)
		self.r_cal = DateEntry(frame, width=12, background='darkblue',
				foreground='white', borderwidth=2,
				date_pattern='dd/mm/yyyy')
		self.r_cal.grid(row=2, column=1)
		self.l_cal = DateEntry(frame, width=12, background='darkblue',
				foreground='white', borderwidth=2,
				date_pattern='dd/mm/yyyy')
		self.l_cal.grid(row=2, column=2)

		self.r_cal.bind("<<DateEntrySelected>>", self.onchange_r_cal)
		self.l_cal.bind("<<DateEntrySelected>>", self.onchange_l_cal)


		tk.Label(frame, text="PROSTHESIS").grid(sticky="W", row=3, column=0)
		# ttk.Entry(frame, width=15).grid(row=3, column=1, padx=5, pady=5)
		# ttk.Entry(frame, width=15).grid(row=3, column=2, padx=5, pady=5)

		
		self.r_pros = ttk.Combobox(frame, width = 12, textvariable = self.sv_r_pros)
		self.l_pros = ttk.Combobox(frame, width = 12, textvariable = self.sv_l_pros)

		# self.r_pros['values'] = self.parent.controller.user_pref_obj['IMPLANT_NAMES']['UKR']
		# self.l_pros['values'] =	self.parent.controller.user_pref_obj['IMPLANT_NAMES']['TKR']

		# self.r_pros['values'] = ('SM','SP','AT','OX','HP','NexGen','Genesis II')
		# self.l_pros['values'] = ('SM','SP','AT','OX','HP','NexGen','Genesis II')
		self.r_pros.grid(row=3, column=1, padx=5, pady=5)
		self.l_pros.grid(row=3, column=2, padx=5, pady=5)


	def constructExcelFrame(self, frame):

		tk.Label(frame, text="MASTER EXCEL EXPORT", font=("TkDefaultFont", 12)).grid(sticky="W", column=0, row=0, columnspan=2)		


		tk.Label(frame, text="UKR PATH").grid(sticky="W", row=1, column=0)
		self.ukr_excel_entry =  ttk.Entry(frame, width=50)
		self.ukr_excel_entry.grid(sticky="W",row=2, column=0)
		ttk.Button(frame, text="SET PATH", command=self.set_ukr_dir).grid(sticky="W",row=3, column=0, pady=5)		


		tk.Label(frame, text="TKR PATH").grid(sticky="W", row=4, column=0, pady=(8,0))
		self.tkr_excel_entry =  ttk.Entry(frame, width=50)
		self.tkr_excel_entry.grid(sticky="W",row=5, column=0)
		ttk.Button(frame, text="SET PATH", command=self.set_tkr_dir).grid(sticky="W",row=6, column=0, pady=5)		


		rb_frame = ttk.Frame(frame)
		rb_frame.grid(sticky="W",row=7, column=0, pady=(10,0))

		rb1 = ttk.Radiobutton(rb_frame, text='UKR', value='UKR', variable=self.sv_utkr)
		rb2 = ttk.Radiobutton(rb_frame, text='TKR', value='TKR', variable=self.sv_utkr)
		rb1.pack(side="left")	
		rb2.pack(side="left")


		self.m_excel_btn = ttk.Button(frame, text="EXPORT TO MASTER", command=self.btn_export_master)
		self.m_excel_btn.grid(sticky="W",row=8, column=0)

		# check if master excel paths exist
		jsonFile = open("user_prefs.json", "r") # Open the JSON file for reading
		prefs = json.load(jsonFile) # Read the JSON into the buffer
		jsonFile.close() # Close the JSON file


		ukr_path = prefs["MASTER_EXCEL_PATHS"]["UKR_PATH"]
		tkr_path = prefs["MASTER_EXCEL_PATHS"]["TKR_PATH"]
		if ukr_path != None:			
			self.ukr_excel_file = ukr_path
			self.ukr_excel_entry.delete(0, tk.END)
			self.ukr_excel_entry.insert(0, ukr_path)

		if tkr_path != None:
			self.tkr_excel_file = tkr_path
			self.tkr_excel_entry.delete(0, tk.END)
			self.tkr_excel_entry.insert(0, tkr_path)



	def set_tkr_dir(self):
		fileobj = filedialog.askopenfile()
		self.tkr_excel_file = fileobj.name		
		# print(type(self.tkr_excel_file))
		# print(self.tkr_excel_file.name)
		if isinstance(self.tkr_excel_file, str) and self.tkr_excel_file != "":
			self.tkr_excel_entry.delete(0, tk.END)
			self.tkr_excel_entry.insert(0, self.tkr_excel_file)

			jsonFile = open("user_prefs.json", "r") # Open the JSON file for reading
			data = json.load(jsonFile) # Read the JSON into the buffer
			jsonFile.close() # Close the JSON file

			# Working with buffered content			
			data["MASTER_EXCEL_PATHS"]["TKR_PATH"] = self.tkr_excel_file			

			# Save our changes to JSON file
			jsonFile = open("user_prefs.json", "w+")
			jsonFile.write(json.dumps(data, ensure_ascii=False, indent=4))
			jsonFile.close()

	def set_ukr_dir(self):
		fileobj = filedialog.askopenfile()
		self.ukr_excel_file = fileobj.name
		# print(type(self.ukr_excel_file))
		# print(self.ukr_excel_file.name)
		if isinstance(self.ukr_excel_file, str) and self.ukr_excel_file != "":
			self.ukr_excel_entry.delete(0, tk.END)
			self.ukr_excel_entry.insert(0, self.ukr_excel_file)

			jsonFile = open("user_prefs.json", "r") # Open the JSON file for reading
			data = json.load(jsonFile) # Read the JSON into the buffer
			jsonFile.close() # Close the JSON file

			# Working with buffered content			
			data["MASTER_EXCEL_PATHS"]["UKR_PATH"] = self.ukr_excel_file

			# Save our changes to JSON file
			jsonFile = open("user_prefs.json", "w+")			
			jsonFile.write(json.dumps(data, ensure_ascii=False, indent=4))
			jsonFile.close()



	def entry_callback(self, entry_type, var):
		print('{} entry_type update {}'.format(entry_type, var.get()))
		val = var.get().strip()
		if val != "":
			if entry_type == "F-NAME":
				self.master_dict["DETAILS"]["F_NAME"] = val
				self.parent.controller.save_json()
			elif entry_type == "L-NAME":
				self.master_dict["DETAILS"]["L_NAME"] = val
				self.parent.controller.save_json()
			elif entry_type == "AGE":
				self.master_dict["DETAILS"]["AGE"] = val
				self.parent.controller.save_json()


			elif entry_type == "R-KL":
				self.master_dict["DETAILS"]["R_KL_GRADE"] = val
				self.parent.controller.save_json()
			elif entry_type == "L-KL":
				self.master_dict["DETAILS"]["L_KL_GRADE"] = val
				self.parent.controller.save_json()


			elif entry_type == "R-PROS":
				self.master_dict["DETAILS"]["R_PROS"] = val
				self.parent.controller.save_json()
			elif entry_type == "L-PROS":
				self.master_dict["DETAILS"]["L_PROS"] = val
				self.parent.controller.save_json()
			
	def radio_callback(self, var):

		sex = var.get().strip()

		if sex != None:
			self.master_dict["DETAILS"]["SEX"] = sex
			self.parent.controller.save_json()

		print('radio callback {}'.format(sex))

	def utkr_callback(self, var):
		utkr = var.get().strip()
		print('utkr callback {}'.format(utkr))

		if utkr == "UKR":
			self.r_pros['values'] = self.parent.controller.user_pref_obj['IMPLANT_NAMES']['UKR']
			self.l_pros['values'] =	self.parent.controller.user_pref_obj['IMPLANT_NAMES']['UKR']
		elif utkr == "TKR":
			self.r_pros['values'] = self.parent.controller.user_pref_obj['IMPLANT_NAMES']['TKR']
			self.l_pros['values'] =	self.parent.controller.user_pref_obj['IMPLANT_NAMES']['TKR']



	def onchange_r_cal(self, event):
		self.master_dict["DETAILS"]["R_SX_DATE"] = self.r_cal.get_date().strftime('%d/%m/%Y')
		self.parent.controller.save_json()

	def onchange_l_cal(self, event):
		self.master_dict["DETAILS"]["L_SX_DATE"] = self.l_cal.get_date().strftime('%d/%m/%Y')
		self.parent.controller.save_json()


	def warningBox(self, message):
		'''Display a warning box with message'''
		messagebox.showwarning("Warning", message)

	def btn_export_master(self):
		uktr_val = self.sv_utkr.get()

		if uktr_val == "":
			self.warningBox("Choose UKR or TKR")
					
		elif uktr_val == "UKR":
			if self.ukr_excel_file == None:
				self.warningBox("No UKR File")
			else:
				# df = self.construct_df("UKR")
				# print(df)
				# print(type(self.ukr_excel_file))
				# self.append_df_to_excel(self.ukr_excel_file, df)
				self.m_excel_btn.configure(state=DISABLED)
				self.queue = queue.Queue()
				ThreadedTask(self.queue, self.master_dict, uktr_val, self.ukr_excel_file).start()
				self.master.after(100, self.process_queue)


		elif uktr_val == "TKR":
			if self.tkr_excel_file == None:
				self.warningBox("No TKR File")
			else:
				df = self.construct_df("TKR")
				# print(df['HKA'])
				# print(type(self.tkr_excel_file))
				# self.append_df_to_excel(self.tkr_excel_file, df)
				self.m_excel_btn.configure(state=DISABLED)
				self.queue = queue.Queue()
				ThreadedTask(self.queue, self.master_dict, uktr_val, self.tkr_excel_file).start()
				self.master.after(100, self.process_queue)

	def process_queue(self):
		try:
			msg = self.queue.get(0)
			# Show result of the task if needed
			print(msg)

			self.m_excel_btn.configure(state=NORMAL)
			if msg == "Possible File open":
				self.warningBox("Close Excel File and try again")
			elif msg == "Task finished":
				self.warningBox("Success")

			# self.prog_bar.stop()
		except queue.Empty:
			self.master.after(100, self.process_queue)


	def construct_df(self, etype=None):
		if etype == None:
			return

		# handle no name information case		
		l_name 		= ""
		f_name 		= ""
		disp_age 	= ""
		disp_sex 	= ""		
	
		# NAME
		if self.master_dict["DETAILS"]["L_NAME"] != None:
			l_name = self.master_dict["DETAILS"]["L_NAME"]
		else:
			print("lname not fpund")
		if self.master_dict["DETAILS"]["F_NAME"] != None:
			f_name = self.master_dict["DETAILS"]["F_NAME"]
		else:
			print("fname not fpund")
		disp_name 	= (l_name + " " + f_name).upper()		

		# AGE
		if self.master_dict["DETAILS"]["AGE"] != None:
			disp_age = self.master_dict["DETAILS"]["AGE"]
		# SEX
		if self.master_dict["DETAILS"]["SEX"] != None:
			disp_sex = self.master_dict["DETAILS"]["SEX"]



		name 		= []
		age 		= []
		gender 		= []
		leg_side 	= []
		k_l_grade 	= []
		mdate 		= []
		prosthesis 	= []
		hka 		= []
		mad 		= []
		mnsa 		= []
		vca 		= []
		afta 		= []
		mldfa 		= []
		aldfa 		= []
		eadfa 		= []
		eadfps 		= []
		eadfds 		= []
		jda 		= []
		tamd 		= []
		mpta 		= []
		kjlo 		= []
		kaol 		= []
		eadta 		= []
		eadtps 		= []
		eadtds 		= []
		fflex 		= []
		pcor 		= []
		acor 		= []
		tslope 		= []
		isr 		= []
		sa 			= []
		ptilt 		= []
		ppba 		= []
		miscell 	= []
		post_hka	= []
		post_mad 	= []
		post_mnsa 	= []
		post_vca 	= []
		post_afta 	= []
		post_mldfa 	= []
		post_aldfa 	= []
		post_eadfa 	= []
		post_eadfps = []
		post_eadfds = []
		post_jda 	= []
		post_tamd 	= []
		post_mpta 	= []
		post_kjlo 	= []
		post_kaol 	= []
		post_eadta 	= []
		post_eadtps = []
		post_eadtds = []
		post_fflex  = []
		post_pcor 	= []
		post_acor 	= []
		post_tslope = []
		post_isr 	= []
		post_sa 	= []
		post_ptilt 	= []
		post_ppba 	= []
		post_miscel = []

		if etype == "UKR":
			fvarval 	= []
			tvarval 	= []
			ffleext 	= []
			tfleext 	= []

		jca 		= []
		lpfa		= []
		mpfa		= []
		ldta		= []
		a_slope 	= []
		p_ma 		= []
		post_jca	= []
		post_lpfa	= []
		post_mpfa	= []
		post_ldta	= []
		post_a_slope = []
		post_p_ma 	= []
		

		if etype == "UKR":
			
			mdict = {
					"NAME"			:	name,
					"AGE"			:	age,
					"GENDER"		:	gender,
					"SIDE"			:	leg_side,
					"K-L GRADE"		:	k_l_grade,
					"DATE"			:	mdate,
					"PROSTHESIS"	:	prosthesis,
					"HKA"			:	hka,
					"MAD"			:	mad,
					"MNSA"			:	mnsa,
					"VCA"			:	vca,
					"aFTA"			:	afta,
					"mLDFA"			:	mldfa,
					"aLDFA"			:	aldfa,
					"EADFA"			:	eadfa,
					"EADFPS"		:	eadfps,
					"EADFDS"		:	eadfds,
					"JDA"			:	jda,
					"TAMD"			:	tamd,
					"MPTA"			:	mpta,
					"KJLO"			:	kjlo,
					"KAOL"			:	kaol,
					"EADTA"			:	eadta,
					"EADTPS"		:	eadtps,
					"EADTDS"		:	eadtds,
					"FFLEX"			:	fflex,
					"PCOR"			:	pcor,
					"ACOR"			:	acor,
					"TSLOPE"		:	tslope,
					"ISR"			:	isr,
					"SA"			:	sa,
					"PTILT"			:	ptilt,
					"PPBA"			:	ppba,
					"MISCELL"		:	miscell,
					"Post-HKA"		:	post_hka,
					"Post-MAD"		:	post_mad,
					"Post-MNSA"		:	post_mnsa,
					"Post-VCA"		:	post_vca,
					"Post-aFTA"		:	post_afta,
					"Post-mLDFA"	:	post_mldfa,
					"Post-aLDFA"	:	post_aldfa,
					"Post-EADFA"	:	post_eadfa,
					"Post-EADFPS"	:	post_eadfps,
					"Post-EADFDS"	:	post_eadfds,
					"Post-JDA"		:	post_jda,
					"Post-TAMD"		:	post_tamd,
					"Post-MPTA"		:	post_mpta,
					"Post-KJLO"		:	post_kjlo,
					"Post-KAOL"		:	post_kaol,
					"Post-EADTA"	:	post_eadta,
					"Post-EADTPS"	:	post_eadtps,
					"Post-EADTDS"	:	post_eadtds,
					"Post-FFLEX	"	:	post_fflex,
					"Post-PCOR"		:	post_pcor,
					"Post-ACOR"		:	post_acor,
					"Post-TSLOPE"	:	post_tslope,
					"Post-ISR"		:	post_isr,
					"Post-SA"		:	post_sa,
					"Post-PTILT"	:	post_ptilt,
					"Post-PPBA"		:	post_ppba,
					"Post-MISCELL"	:	post_miscel,
					"Post-F FLE/EXT"	:	ffleext,
					"Post-F VAR/VAL"	:	fvarval,
					"Post-T FLE/EXT"	:	tfleext,
					"Post-T VAR/VAL"	:	tvarval,
					"JCA"			:	jca,
					"LPFA"			:	lpfa,
					"MPFA"			:	mpfa,
					"LDTA"			:	ldta,
					"ANKLE_SLOPE"	:	a_slope,
					"pMA"			:	p_ma,
					"Post-JCA"		:	post_jca,
					"Post-LPFA"		:	post_lpfa,
					"Post-MPFA"		:	post_mpfa,
					"Post-LDTA"		:	post_ldta,
					"Post-ANKLE_SLOPE"	:	post_a_slope,
					"Post-pMA"		: 	post_p_ma					
				}

		else:
			mdict = {
					"NAME"			:	name,
					"AGE"			:	age,
					"GENDER"		:	gender,
					"SIDE"			:	leg_side,
					"K-L GRADE"		:	k_l_grade,
					"DATE"			:	mdate,
					"PROSTHESIS"	:	prosthesis,
					"HKA"			:	hka,
					"MAD"			:	mad,
					"MNSA"			:	mnsa,
					"VCA"			:	vca,
					"aFTA"			:	afta,
					"mLDFA"			:	mldfa,
					"aLDFA"			:	aldfa,
					"EADFA"			:	eadfa,
					"EADFPS"		:	eadfps,
					"EADFDS"		:	eadfds,
					"JDA"			:	jda,
					"TAMD"			:	tamd,
					"MPTA"			:	mpta,
					"KJLO"			:	kjlo,
					"KAOL"			:	kaol,
					"EADTA"			:	eadta,
					"EADTPS"		:	eadtps,
					"EADTDS"		:	eadtds,
					"FFLEX"			:	fflex,
					"PCOR"			:	pcor,
					"ACOR"			:	acor,
					"TSLOPE"		:	tslope,
					"ISR"			:	isr,
					"SA"			:	sa,
					"PTILT"			:	ptilt,
					"PPBA"			:	ppba,
					"MISCELL"		:	miscell,
					"Post-HKA"		:	post_hka,
					"Post-MAD"		:	post_mad,
					"Post-MNSA"		:	post_mnsa,
					"Post-VCA"		:	post_vca,
					"Post-aFTA"		:	post_afta,
					"Post-mLDFA"	:	post_mldfa,
					"Post-aLDFA"	:	post_aldfa,
					"Post-EADFA"	:	post_eadfa,
					"Post-EADFPS"	:	post_eadfps,
					"Post-EADFDS"	:	post_eadfds,
					"Post-JDA"		:	post_jda,
					"Post-TAMD"		:	post_tamd,
					"Post-MPTA"		:	post_mpta,
					"Post-KJLO"		:	post_kjlo,
					"Post-KAOL"		:	post_kaol,
					"Post-EADTA"	:	post_eadta,
					"Post-EADTPS"	:	post_eadtps,
					"Post-EADTDS"	:	post_eadtds,
					"Post-FFLEX	"	:	post_fflex,
					"Post-PCOR"		:	post_pcor,
					"Post-ACOR"		:	post_acor,
					"Post-TSLOPE"	:	post_tslope,
					"Post-ISR"		:	post_isr,
					"Post-SA"		:	post_sa,
					"Post-PTILT"	:	post_ptilt,
					"Post-PPBA"		:	post_ppba,
					"Post-MISCELL"	:	post_miscel,					
					"JCA"			:	jca,
					"LPFA"			:	lpfa,
					"MPFA"			:	mpfa,
					"LDTA"			:	ldta,
					"ANKLE_SLOPE"	:	a_slope,
					"pMA"			:	p_ma,
					"Post-JCA"		:	post_jca,
					"Post-LPFA"		:	post_lpfa,
					"Post-MPFA"		:	post_mpfa,
					"Post-LDTA"		:	post_ldta,
					"Post-ANKLE_SLOPE"	:	post_a_slope,
					"Post-pMA"		: 	post_p_ma					
				}





		# Iterate
		# for op_type in ["PRE-OP", "POST-OP"]:
		for side in ["RIGHT", "LEFT"]:


			x = ""

			# master list				
			if self.master_dict["EXCEL"]["PRE-OP"][side]["HASDATA"] == True:
				name.append(disp_name)
				age.append(disp_age)
				gender.append(disp_sex)
				leg_side.append(side[0])
				if side == "RIGHT":
					k_l_grade.append(self.master_dict["DETAILS"]["R_KL_GRADE"])
					prosthesis.append(self.master_dict["DETAILS"]["R_PROS"])
					mdate.append(self.master_dict["DETAILS"]["R_SX_DATE"])
					
				elif side == "LEFT":
					k_l_grade.append(self.master_dict["DETAILS"]["L_KL_GRADE"])
					prosthesis.append(self.master_dict["DETAILS"]["L_PROS"])
					mdate.append(self.master_dict["DETAILS"]["L_SX_DATE"])
				
								
				hka.append(self.master_dict["EXCEL"]["PRE-OP"][side]["HKA"])
				mad.append(self.master_dict["EXCEL"]["PRE-OP"][side]["MAD"])
				mnsa.append(self.master_dict["EXCEL"]["PRE-OP"][side]["MNSA"])
				vca.append(self.master_dict["EXCEL"]["PRE-OP"][side]["VCA"])
				afta.append(self.master_dict["EXCEL"]["PRE-OP"][side]["aFTA"])
				mldfa.append(self.master_dict["EXCEL"]["PRE-OP"][side]["mLDFA"])
				aldfa.append(self.master_dict["EXCEL"]["PRE-OP"][side]["aLDFA"])
				eadfa.append(self.master_dict["EXCEL"]["PRE-OP"][side]["EADFA"])
				eadfps.append(self.master_dict["EXCEL"]["PRE-OP"][side]["EADFPS"])
				eadfds.append(self.master_dict["EXCEL"]["PRE-OP"][side]["EADFDS"])
				jda.append(self.master_dict["EXCEL"]["PRE-OP"][side]["JDA"])
				tamd.append(self.master_dict["EXCEL"]["PRE-OP"][side]["TAMD"])
				mpta.append(self.master_dict["EXCEL"]["PRE-OP"][side]["MPTA"])
				kjlo.append(self.master_dict["EXCEL"]["PRE-OP"][side]["KJLO"])
				kaol.append(self.master_dict["EXCEL"]["PRE-OP"][side]["KAOL"])
				eadta.append(self.master_dict["EXCEL"]["PRE-OP"][side]["EADTA"])
				eadtps.append(self.master_dict["EXCEL"]["PRE-OP"][side]["EADTPS"])
				eadtds.append(self.master_dict["EXCEL"]["PRE-OP"][side]["EADTDS"])
				fflex.append(self.master_dict["EXCEL"]["PRE-OP"][side]["FFLEX"])
				pcor.append(self.master_dict["EXCEL"]["PRE-OP"][side]["PCOR"])
				acor.append(self.master_dict["EXCEL"]["PRE-OP"][side]["ACOR"])
				tslope.append(self.master_dict["EXCEL"]["PRE-OP"][side]["TSLOPE"])
				isr.append(self.master_dict["EXCEL"]["PRE-OP"][side]["ISR"])
				sa.append(self.master_dict["EXCEL"]["PRE-OP"][side]["SA"])
				ptilt.append(self.master_dict["EXCEL"]["PRE-OP"][side]["PTILT"])
				ppba.append(self.master_dict["EXCEL"]["PRE-OP"][side]["PPBA"])
				miscell.append(x)

						
				post_hka.append(self.master_dict["EXCEL"]["POST-OP"][side]["HKA"])
				post_mad.append(self.master_dict["EXCEL"]["POST-OP"][side]["MAD"])
				post_mnsa.append(self.master_dict["EXCEL"]["POST-OP"][side]["MNSA"])
				post_vca.append(self.master_dict["EXCEL"]["POST-OP"][side]["VCA"])
				post_afta.append(self.master_dict["EXCEL"]["POST-OP"][side]["aFTA"])
				post_mldfa.append(self.master_dict["EXCEL"]["POST-OP"][side]["mLDFA"])
				post_aldfa.append(self.master_dict["EXCEL"]["POST-OP"][side]["aLDFA"])
				post_eadfa.append(self.master_dict["EXCEL"]["POST-OP"][side]["EADFA"])
				post_eadfps.append(self.master_dict["EXCEL"]["POST-OP"][side]["EADFPS"])
				post_eadfds.append(self.master_dict["EXCEL"]["POST-OP"][side]["EADFDS"])
				post_jda.append(self.master_dict["EXCEL"]["POST-OP"][side]["JDA"])
				post_tamd.append(self.master_dict["EXCEL"]["POST-OP"][side]["TAMD"])
				post_mpta.append(self.master_dict["EXCEL"]["POST-OP"][side]["MPTA"])
				post_kjlo.append(self.master_dict["EXCEL"]["POST-OP"][side]["KJLO"])
				post_kaol.append(self.master_dict["EXCEL"]["POST-OP"][side]["KAOL"])
				post_eadta.append(self.master_dict["EXCEL"]["POST-OP"][side]["EADTA"])
				post_eadtps.append(self.master_dict["EXCEL"]["POST-OP"][side]["EADTPS"])
				post_eadtds.append(self.master_dict["EXCEL"]["POST-OP"][side]["EADTDS"])
				post_fflex.append(self.master_dict["EXCEL"]["POST-OP"][side]["FFLEX"])
				post_pcor.append(self.master_dict["EXCEL"]["POST-OP"][side]["PCOR"])
				post_acor.append(self.master_dict["EXCEL"]["POST-OP"][side]["ACOR"])
				post_tslope.append(self.master_dict["EXCEL"]["POST-OP"][side]["TSLOPE"])
				post_isr.append(self.master_dict["EXCEL"]["POST-OP"][side]["ISR"])
				post_sa.append(self.master_dict["EXCEL"]["POST-OP"][side]["SA"])
				post_ptilt.append(self.master_dict["EXCEL"]["POST-OP"][side]["PTILT"])
				post_ppba.append(self.master_dict["EXCEL"]["POST-OP"][side]["PPBA"])
				post_miscel.append(x)

				if etype == "UKR":
					fvarval.append(self.master_dict["EXCEL"]["POST-OP"][side]["FVAR/VAL"])
					tvarval.append(self.master_dict["EXCEL"]["POST-OP"][side]["TVAR/VAL"])
					ffleext.append(self.master_dict["EXCEL"]["POST-OP"][side]["FFLE/EXT"])
					tfleext.append(self.master_dict["EXCEL"]["POST-OP"][side]["TFLE/EXT"])

				jca.append(self.master_dict["EXCEL"]["PRE-OP"][side]["JCA"])
				lpfa.append(self.master_dict["EXCEL"]["PRE-OP"][side]["LPFA"])
				mpfa.append(self.master_dict["EXCEL"]["PRE-OP"][side]["MPFA"])
				ldta.append(self.master_dict["EXCEL"]["PRE-OP"][side]["LDTA"])
				a_slope.append(self.master_dict["EXCEL"]["PRE-OP"][side]["ANKLE_SLOPE"])
				p_ma.append(self.master_dict["EXCEL"]["PRE-OP"][side]["pMA"])
				post_jca.append(self.master_dict["EXCEL"]["POST-OP"][side]["JCA"])
				post_lpfa.append(self.master_dict["EXCEL"]["POST-OP"][side]["LPFA"])
				post_mpfa.append(self.master_dict["EXCEL"]["POST-OP"][side]["MPFA"])
				post_ldta.append(self.master_dict["EXCEL"]["POST-OP"][side]["LDTA"])
				post_a_slope.append(self.master_dict["EXCEL"]["POST-OP"][side]["ANKLE_SLOPE"])
				post_p_ma.append(self.master_dict["EXCEL"]["POST-OP"][side]["pMA"])

		# print(mtype)
		# import pprint
		# pprint.pprint(mdict)
		df = pd.DataFrame(mdict)

		return df


	def append_df_to_excel(self, filename, df, sheet_name='Sheet1', startrow=None,
					   truncate_sheet=False, 
					   **to_excel_kwargs):
		"""
		Append a DataFrame [df] to existing Excel file [filename]
		into [sheet_name] Sheet.
		If [filename] doesn't exist, then this function will create it.

		@param filename: File path or existing ExcelWriter
						 (Example: '/path/to/file.xlsx')
		@param df: DataFrame to save to workbook
		@param sheet_name: Name of sheet which will contain DataFrame.
						   (default: 'Sheet1')
		@param startrow: upper left cell row to dump data frame.
						 Per default (startrow=None) calculate the last row
						 in the existing DF and write to the next row...
		@param truncate_sheet: truncate (remove and recreate) [sheet_name]
							   before writing DataFrame to Excel file
		@param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
								[can be a dictionary]
		@return: None

		Usage examples:

		>>> append_df_to_excel('d:/temp/test.xlsx', df)

		>>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

		>>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
							   index=False)

		>>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2', 
							   index=False, startrow=25)

		(c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
		"""
		# Excel file doesn't exist - saving and exiting		
		# if not Path(filename).is_file():
		if not os.path.isfile(filename):
			df.to_excel(
				filename,
				sheet_name=sheet_name, 
				startrow=startrow if startrow is not None else 0, 
				**to_excel_kwargs)
			return
		
		# ignore [engine] parameter if it was passed
		if 'engine' in to_excel_kwargs:
			to_excel_kwargs.pop('engine')

		writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

		# try to open an existing workbook
		writer.book = load_workbook(filename)
		
		# get the last row in the existing Excel sheet
		# if it was not specified explicitly
		if startrow is None and sheet_name in writer.book.sheetnames:
			# startrow = writer.book[sheet_name].max_row
			# startrow = writer.book.active.max_row				
			startrow = len(list(writer.book.active.rows))
			val = writer.book.active.cell(startrow,1).value
			if val == None:
				for x in range(startrow, 0, -1):
					# print(x)
					check = writer.book.active.cell(x,1).value
					print(check)
					if check != None:
						break
				startrow = x


		# truncate sheet
		if truncate_sheet and sheet_name in writer.book.sheetnames:
			# index of [sheet_name] sheet
			idx = writer.book.sheetnames.index(sheet_name)
			# remove [sheet_name]
			writer.book.remove(writer.book.worksheets[idx])
			# create an empty sheet [sheet_name] using old index
			writer.book.create_sheet(sheet_name, idx)
		
		# copy existing sheets
		writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

		if startrow is None:
			startrow = 0

		# write out the new sheet
		# df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
		df.to_excel(writer, sheet_name, startrow=startrow, index=False, header=False, **to_excel_kwargs)	

		# save the workbook
		writer.save()


class ThreadedTask(threading.Thread):
	def __init__(self, queue, master_dict, utkr_val, excel_path):
		threading.Thread.__init__(self)
		self.queue = queue
		self.master_dict = master_dict
		self.utkr = utkr_val
		self.excel_path = excel_path
	def run(self):
		# time.sleep(5)  # Simulate long running process

		df = self.construct_df(self.utkr)
		# print(df)
		# print(list(df))

		
		# cols = df.columns.drop('NAME')
		# cols = df.columns.drop('GENDER')
		# cols = df.columns.drop('SIDE')
		# cols = df.columns.drop('DATE')
		# cols = df.columns.drop('PROSTHESIS')
		# cols = df.columns.drop('MISCELL')
		# cols = df.columns.drop('Post-MISCELL')
		# cols = df.columns.drop('Post-pMA')

		# df[cols] = df[cols].apply(pd.to_numeric)
		df = df.apply(pd.to_numeric, errors='ignore')

		print(df)
		# print(type(self.ukr_excel_file))
		try:
			self.append_df_to_excel(self.excel_path, df)
			self.queue.put("Task finished")
		except Exception as e:
			# raise e
			print(e)
			self.queue.put("Possible File open")
		

		# self.queue.put("Task finished")


	def construct_df(self, etype=None):
		if etype == None:
			return

		# handle no name information case		
		l_name 		= ""
		f_name 		= ""
		disp_age 	= ""
		disp_sex 	= ""		
	
		# NAME
		if self.master_dict["DETAILS"]["L_NAME"] != None:
			l_name = self.master_dict["DETAILS"]["L_NAME"]
		else:
			print("lname not fpund")
		if self.master_dict["DETAILS"]["F_NAME"] != None:
			f_name = self.master_dict["DETAILS"]["F_NAME"]
		else:
			print("fname not fpund")
		disp_name 	= (l_name + " " + f_name).upper()		

		# AGE
		if self.master_dict["DETAILS"]["AGE"] != None:
			disp_age = self.master_dict["DETAILS"]["AGE"]
		# SEX
		if self.master_dict["DETAILS"]["SEX"] != None:
			disp_sex = self.master_dict["DETAILS"]["SEX"]



		name 		= []
		age 		= []
		gender 		= []
		leg_side 	= []
		k_l_grade 	= []
		mdate 		= []
		prosthesis 	= []
		hka 		= []
		mad 		= []
		mnsa 		= []
		vca 		= []
		afta 		= []
		mldfa 		= []
		aldfa 		= []
		eadfa 		= []
		eadfps 		= []
		eadfds 		= []
		jda 		= []
		tamd 		= []
		mpta 		= []
		kjlo 		= []
		kaol 		= []
		eadta 		= []
		eadtps 		= []
		eadtds 		= []
		fflex 		= []
		pcor 		= []
		acor 		= []
		tslope 		= []
		isr 		= []
		sa 			= []
		ptilt 		= []
		ppba 		= []
		miscell 	= []
		post_hka	= []
		post_mad 	= []
		post_mnsa 	= []
		post_vca 	= []
		post_afta 	= []
		post_mldfa 	= []
		post_aldfa 	= []
		post_eadfa 	= []
		post_eadfps = []
		post_eadfds = []
		post_jda 	= []
		post_tamd 	= []
		post_mpta 	= []
		post_kjlo 	= []
		post_kaol 	= []
		post_eadta 	= []
		post_eadtps = []
		post_eadtds = []
		post_fflex  = []
		post_pcor 	= []
		post_acor 	= []
		post_tslope = []
		post_isr 	= []
		post_sa 	= []
		post_ptilt 	= []
		post_ppba 	= []
		post_miscel = []

		if etype == "UKR":
			fvarval 	= []
			tvarval 	= []
			ffleext 	= []
			tfleext 	= []

		jca 		= []
		lpfa		= []
		mpfa		= []
		ldta		= []
		a_slope 	= []
		p_ma 		= []
		post_jca	= []
		post_lpfa	= []
		post_mpfa	= []
		post_ldta	= []
		post_a_slope = []
		post_p_ma 	= []
		

		if etype == "UKR":
			
			mdict = {
					"NAME"			:	name,
					"AGE"			:	age,
					"GENDER"		:	gender,
					"SIDE"			:	leg_side,
					"K-L GRADE"		:	k_l_grade,
					"DATE"			:	mdate,
					"PROSTHESIS"	:	prosthesis,
					"HKA"			:	hka,
					"MAD"			:	mad,
					"MNSA"			:	mnsa,
					"VCA"			:	vca,
					"aFTA"			:	afta,
					"mLDFA"			:	mldfa,
					"aLDFA"			:	aldfa,
					"EADFA"			:	eadfa,
					"EADFPS"		:	eadfps,
					"EADFDS"		:	eadfds,
					"JDA"			:	jda,
					"TAMD"			:	tamd,
					"MPTA"			:	mpta,
					"KJLO"			:	kjlo,
					"KAOL"			:	kaol,
					"EADTA"			:	eadta,
					"EADTPS"		:	eadtps,
					"EADTDS"		:	eadtds,
					"FFLEX"			:	fflex,
					"PCOR"			:	pcor,
					"ACOR"			:	acor,
					"TSLOPE"		:	tslope,
					"ISR"			:	isr,
					"SA"			:	sa,
					"PTILT"			:	ptilt,
					"PPBA"			:	ppba,
					"MISCELL"		:	miscell,
					"Post-HKA"		:	post_hka,
					"Post-MAD"		:	post_mad,
					"Post-MNSA"		:	post_mnsa,
					"Post-VCA"		:	post_vca,
					"Post-aFTA"		:	post_afta,
					"Post-mLDFA"	:	post_mldfa,
					"Post-aLDFA"	:	post_aldfa,
					"Post-EADFA"	:	post_eadfa,
					"Post-EADFPS"	:	post_eadfps,
					"Post-EADFDS"	:	post_eadfds,
					"Post-JDA"		:	post_jda,
					"Post-TAMD"		:	post_tamd,
					"Post-MPTA"		:	post_mpta,
					"Post-KJLO"		:	post_kjlo,
					"Post-KAOL"		:	post_kaol,
					"Post-EADTA"	:	post_eadta,
					"Post-EADTPS"	:	post_eadtps,
					"Post-EADTDS"	:	post_eadtds,
					"Post-FFLEX	"	:	post_fflex,
					"Post-PCOR"		:	post_pcor,
					"Post-ACOR"		:	post_acor,
					"Post-TSLOPE"	:	post_tslope,
					"Post-ISR"		:	post_isr,
					"Post-SA"		:	post_sa,
					"Post-PTILT"	:	post_ptilt,
					"Post-PPBA"		:	post_ppba,
					"Post-MISCELL"	:	post_miscel,
					"Post-F FLE/EXT"	:	ffleext,
					"Post-F VAR/VAL"	:	fvarval,
					"Post-T FLE/EXT"	:	tfleext,
					"Post-T VAR/VAL"	:	tvarval,
					"JCA"			:	jca,
					"LPFA"			:	lpfa,
					"MPFA"			:	mpfa,
					"LDTA"			:	ldta,
					"ANKLE_SLOPE"	:	a_slope,
					"pMA"			:	p_ma,
					"Post-JCA"		:	post_jca,
					"Post-LPFA"		:	post_lpfa,
					"Post-MPFA"		:	post_mpfa,
					"Post-LDTA"		:	post_ldta,
					"Post-ANKLE_SLOPE"	:	post_a_slope,
					"Post-pMA"		: 	post_p_ma					
				}

		else:
			mdict = {
					"NAME"			:	name,
					"AGE"			:	age,
					"GENDER"		:	gender,
					"SIDE"			:	leg_side,
					"K-L GRADE"		:	k_l_grade,
					"DATE"			:	mdate,
					"PROSTHESIS"	:	prosthesis,
					"HKA"			:	hka,
					"MAD"			:	mad,
					"MNSA"			:	mnsa,
					"VCA"			:	vca,
					"aFTA"			:	afta,
					"mLDFA"			:	mldfa,
					"aLDFA"			:	aldfa,
					"EADFA"			:	eadfa,
					"EADFPS"		:	eadfps,
					"EADFDS"		:	eadfds,
					"JDA"			:	jda,
					"TAMD"			:	tamd,
					"MPTA"			:	mpta,
					"KJLO"			:	kjlo,
					"KAOL"			:	kaol,
					"EADTA"			:	eadta,
					"EADTPS"		:	eadtps,
					"EADTDS"		:	eadtds,
					"FFLEX"			:	fflex,
					"PCOR"			:	pcor,
					"ACOR"			:	acor,
					"TSLOPE"		:	tslope,
					"ISR"			:	isr,
					"SA"			:	sa,
					"PTILT"			:	ptilt,
					"PPBA"			:	ppba,
					"MISCELL"		:	miscell,
					"Post-HKA"		:	post_hka,
					"Post-MAD"		:	post_mad,
					"Post-MNSA"		:	post_mnsa,
					"Post-VCA"		:	post_vca,
					"Post-aFTA"		:	post_afta,
					"Post-mLDFA"	:	post_mldfa,
					"Post-aLDFA"	:	post_aldfa,
					"Post-EADFA"	:	post_eadfa,
					"Post-EADFPS"	:	post_eadfps,
					"Post-EADFDS"	:	post_eadfds,
					"Post-JDA"		:	post_jda,
					"Post-TAMD"		:	post_tamd,
					"Post-MPTA"		:	post_mpta,
					"Post-KJLO"		:	post_kjlo,
					"Post-KAOL"		:	post_kaol,
					"Post-EADTA"	:	post_eadta,
					"Post-EADTPS"	:	post_eadtps,
					"Post-EADTDS"	:	post_eadtds,
					"Post-FFLEX	"	:	post_fflex,
					"Post-PCOR"		:	post_pcor,
					"Post-ACOR"		:	post_acor,
					"Post-TSLOPE"	:	post_tslope,
					"Post-ISR"		:	post_isr,
					"Post-SA"		:	post_sa,
					"Post-PTILT"	:	post_ptilt,
					"Post-PPBA"		:	post_ppba,
					"Post-MISCELL"	:	post_miscel,					
					"JCA"			:	jca,
					"LPFA"			:	lpfa,
					"MPFA"			:	mpfa,
					"LDTA"			:	ldta,
					"ANKLE_SLOPE"	:	a_slope,
					"pMA"			:	p_ma,
					"Post-JCA"		:	post_jca,
					"Post-LPFA"		:	post_lpfa,
					"Post-MPFA"		:	post_mpfa,
					"Post-LDTA"		:	post_ldta,
					"Post-ANKLE_SLOPE"	:	post_a_slope,
					"Post-pMA"		: 	post_p_ma					
				}





		# Iterate
		# for op_type in ["PRE-OP", "POST-OP"]:
		for side in ["RIGHT", "LEFT"]:


			x = ""

			# master list				
			if self.master_dict["EXCEL"]["PRE-OP"][side]["HASDATA"] == True:
				name.append(disp_name)
				age.append(disp_age)
				gender.append(disp_sex)
				leg_side.append(side[0])
				if side == "RIGHT":
					k_l_grade.append(self.master_dict["DETAILS"]["R_KL_GRADE"])
					prosthesis.append(self.master_dict["DETAILS"]["R_PROS"])
					mdate.append(self.master_dict["DETAILS"]["R_SX_DATE"])
					
				elif side == "LEFT":
					k_l_grade.append(self.master_dict["DETAILS"]["L_KL_GRADE"])
					prosthesis.append(self.master_dict["DETAILS"]["L_PROS"])
					mdate.append(self.master_dict["DETAILS"]["L_SX_DATE"])
				
								
				hka.append(self.master_dict["EXCEL"]["PRE-OP"][side]["HKA"])
				mad.append(self.master_dict["EXCEL"]["PRE-OP"][side]["MAD"])
				mnsa.append(self.master_dict["EXCEL"]["PRE-OP"][side]["MNSA"])
				vca.append(self.master_dict["EXCEL"]["PRE-OP"][side]["VCA"])
				afta.append(self.master_dict["EXCEL"]["PRE-OP"][side]["aFTA"])
				mldfa.append(self.master_dict["EXCEL"]["PRE-OP"][side]["mLDFA"])
				aldfa.append(self.master_dict["EXCEL"]["PRE-OP"][side]["aLDFA"])
				eadfa.append(self.master_dict["EXCEL"]["PRE-OP"][side]["EADFA"])
				eadfps.append(self.master_dict["EXCEL"]["PRE-OP"][side]["EADFPS"])
				eadfds.append(self.master_dict["EXCEL"]["PRE-OP"][side]["EADFDS"])
				jda.append(self.master_dict["EXCEL"]["PRE-OP"][side]["JDA"])
				tamd.append(self.master_dict["EXCEL"]["PRE-OP"][side]["TAMD"])
				mpta.append(self.master_dict["EXCEL"]["PRE-OP"][side]["MPTA"])
				kjlo.append(self.master_dict["EXCEL"]["PRE-OP"][side]["KJLO"])
				kaol.append(self.master_dict["EXCEL"]["PRE-OP"][side]["KAOL"])
				eadta.append(self.master_dict["EXCEL"]["PRE-OP"][side]["EADTA"])
				eadtps.append(self.master_dict["EXCEL"]["PRE-OP"][side]["EADTPS"])
				eadtds.append(self.master_dict["EXCEL"]["PRE-OP"][side]["EADTDS"])
				fflex.append(self.master_dict["EXCEL"]["PRE-OP"][side]["FFLEX"])
				pcor.append(self.master_dict["EXCEL"]["PRE-OP"][side]["PCOR"])
				acor.append(self.master_dict["EXCEL"]["PRE-OP"][side]["ACOR"])
				tslope.append(self.master_dict["EXCEL"]["PRE-OP"][side]["TSLOPE"])
				isr.append(self.master_dict["EXCEL"]["PRE-OP"][side]["ISR"])
				sa.append(self.master_dict["EXCEL"]["PRE-OP"][side]["SA"])
				ptilt.append(self.master_dict["EXCEL"]["PRE-OP"][side]["PTILT"])
				ppba.append(self.master_dict["EXCEL"]["PRE-OP"][side]["PPBA"])
				miscell.append(x)

						
				post_hka.append(self.master_dict["EXCEL"]["POST-OP"][side]["HKA"])
				post_mad.append(self.master_dict["EXCEL"]["POST-OP"][side]["MAD"])
				post_mnsa.append(self.master_dict["EXCEL"]["POST-OP"][side]["MNSA"])
				post_vca.append(self.master_dict["EXCEL"]["POST-OP"][side]["VCA"])
				post_afta.append(self.master_dict["EXCEL"]["POST-OP"][side]["aFTA"])
				post_mldfa.append(self.master_dict["EXCEL"]["POST-OP"][side]["mLDFA"])
				post_aldfa.append(self.master_dict["EXCEL"]["POST-OP"][side]["aLDFA"])
				post_eadfa.append(self.master_dict["EXCEL"]["POST-OP"][side]["EADFA"])
				post_eadfps.append(self.master_dict["EXCEL"]["POST-OP"][side]["EADFPS"])
				post_eadfds.append(self.master_dict["EXCEL"]["POST-OP"][side]["EADFDS"])
				post_jda.append(self.master_dict["EXCEL"]["POST-OP"][side]["JDA"])
				post_tamd.append(self.master_dict["EXCEL"]["POST-OP"][side]["TAMD"])
				post_mpta.append(self.master_dict["EXCEL"]["POST-OP"][side]["MPTA"])
				post_kjlo.append(self.master_dict["EXCEL"]["POST-OP"][side]["KJLO"])
				post_kaol.append(self.master_dict["EXCEL"]["POST-OP"][side]["KAOL"])
				post_eadta.append(self.master_dict["EXCEL"]["POST-OP"][side]["EADTA"])
				post_eadtps.append(self.master_dict["EXCEL"]["POST-OP"][side]["EADTPS"])
				post_eadtds.append(self.master_dict["EXCEL"]["POST-OP"][side]["EADTDS"])
				post_fflex.append(self.master_dict["EXCEL"]["POST-OP"][side]["FFLEX"])
				post_pcor.append(self.master_dict["EXCEL"]["POST-OP"][side]["PCOR"])
				post_acor.append(self.master_dict["EXCEL"]["POST-OP"][side]["ACOR"])
				post_tslope.append(self.master_dict["EXCEL"]["POST-OP"][side]["TSLOPE"])
				post_isr.append(self.master_dict["EXCEL"]["POST-OP"][side]["ISR"])
				post_sa.append(self.master_dict["EXCEL"]["POST-OP"][side]["SA"])
				post_ptilt.append(self.master_dict["EXCEL"]["POST-OP"][side]["PTILT"])
				post_ppba.append(self.master_dict["EXCEL"]["POST-OP"][side]["PPBA"])
				post_miscel.append(x)

				if etype == "UKR":
					fvarval.append(self.master_dict["EXCEL"]["POST-OP"][side]["FVAR/VAL"])
					tvarval.append(self.master_dict["EXCEL"]["POST-OP"][side]["TVAR/VAL"])
					ffleext.append(self.master_dict["EXCEL"]["POST-OP"][side]["FFLE/EXT"])
					tfleext.append(self.master_dict["EXCEL"]["POST-OP"][side]["TFLE/EXT"])

				jca.append(self.master_dict["EXCEL"]["PRE-OP"][side]["JCA"])
				lpfa.append(self.master_dict["EXCEL"]["PRE-OP"][side]["LPFA"])
				mpfa.append(self.master_dict["EXCEL"]["PRE-OP"][side]["MPFA"])
				ldta.append(self.master_dict["EXCEL"]["PRE-OP"][side]["LDTA"])
				a_slope.append(self.master_dict["EXCEL"]["PRE-OP"][side]["ANKLE_SLOPE"])
				p_ma.append(self.master_dict["EXCEL"]["PRE-OP"][side]["pMA"])
				post_jca.append(self.master_dict["EXCEL"]["POST-OP"][side]["JCA"])
				post_lpfa.append(self.master_dict["EXCEL"]["POST-OP"][side]["LPFA"])
				post_mpfa.append(self.master_dict["EXCEL"]["POST-OP"][side]["MPFA"])
				post_ldta.append(self.master_dict["EXCEL"]["POST-OP"][side]["LDTA"])
				post_a_slope.append(self.master_dict["EXCEL"]["POST-OP"][side]["ANKLE_SLOPE"])
				post_p_ma.append(self.master_dict["EXCEL"]["POST-OP"][side]["pMA"])

		# print(mtype)
		# import pprint
		# pprint.pprint(mdict)
		df = pd.DataFrame(mdict)

		return df


	def append_df_to_excel(self, filename, df, sheet_name='Sheet1', startrow=None,
					   truncate_sheet=False, 
					   **to_excel_kwargs):
		"""
		Append a DataFrame [df] to existing Excel file [filename]
		into [sheet_name] Sheet.
		If [filename] doesn't exist, then this function will create it.

		@param filename: File path or existing ExcelWriter
						 (Example: '/path/to/file.xlsx')
		@param df: DataFrame to save to workbook
		@param sheet_name: Name of sheet which will contain DataFrame.
						   (default: 'Sheet1')
		@param startrow: upper left cell row to dump data frame.
						 Per default (startrow=None) calculate the last row
						 in the existing DF and write to the next row...
		@param truncate_sheet: truncate (remove and recreate) [sheet_name]
							   before writing DataFrame to Excel file
		@param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
								[can be a dictionary]
		@return: None

		Usage examples:

		>>> append_df_to_excel('d:/temp/test.xlsx', df)

		>>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

		>>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
							   index=False)

		>>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2', 
							   index=False, startrow=25)

		(c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
		"""
		# Excel file doesn't exist - saving and exiting		
		# if not Path(filename).is_file():
		if not os.path.isfile(filename):
			df.to_excel(
				filename,
				sheet_name=sheet_name, 
				startrow=startrow if startrow is not None else 0, 
				**to_excel_kwargs)
			return
		
		# ignore [engine] parameter if it was passed
		if 'engine' in to_excel_kwargs:
			to_excel_kwargs.pop('engine')

		writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

		# try to open an existing workbook
		writer.book = load_workbook(filename)
		
		# get the last row in the existing Excel sheet
		# if it was not specified explicitly
		if startrow is None and sheet_name in writer.book.sheetnames:
			# startrow = writer.book[sheet_name].max_row
			# startrow = writer.book.active.max_row				
			startrow = len(list(writer.book.active.rows))
			val = writer.book.active.cell(startrow,1).value
			if val == None:
				for x in range(startrow, 0, -1):
					# print(x)
					check = writer.book.active.cell(x,1).value
					print(check)
					if check != None:
						break
				startrow = x


		# truncate sheet
		if truncate_sheet and sheet_name in writer.book.sheetnames:
			# index of [sheet_name] sheet
			idx = writer.book.sheetnames.index(sheet_name)
			# remove [sheet_name]
			writer.book.remove(writer.book.worksheets[idx])
			# create an empty sheet [sheet_name] using old index
			writer.book.create_sheet(sheet_name, idx)
		
		# copy existing sheets
		writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

		if startrow is None:
			startrow = 0

		# write out the new sheet
		# df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
		df.to_excel(writer, sheet_name, startrow=startrow, index=False, header=False, **to_excel_kwargs)	

		# save the workbook
		writer.save()
	


