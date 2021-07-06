
from openpyxl import load_workbook


# writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')
# writer.book = load_workbook(filename)
wb = load_workbook(filename = 'C:/Users/jiggy/Desktop/master_excel/UKR_inbet.xlsx')
ws = wb.active
first_row = ws[1]

# print(first_row)
for cell in first_row:
	print(cell.value)